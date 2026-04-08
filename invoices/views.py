"""
Invoice Management API Views
"""

from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.http import FileResponse, HttpResponse
from django.db.models import Q, Sum, Count, Prefetch
from datetime import datetime, timedelta

from .models import (
    InvoiceSettings,
    Invoice,
    InvoiceLineItem,
    InvoicePayment,
    InvoiceTemplate
)
from .serializers import (
    InvoiceSettingsSerializer,
    InvoiceListSerializer,
    InvoiceDetailSerializer,
    InvoiceCreateUpdateSerializer,
    InvoiceLineItemSerializer,
    InvoicePaymentSerializer,
    InvoiceTemplateSerializer,
    ClientSerializer,
    ListItemMaterialSerializer
)
from ecommerce_crm.models import EcommerceClient, Product
from tickets.models import ItemList, ListItem
from tenants.permissions import require_subscription_feature
from .permissions import CanManageInvoices


class InvoiceSettingsViewSet(viewsets.ModelViewSet):
    """
    API endpoint for invoice settings management
    """
    queryset = InvoiceSettings.objects.select_related('client_itemlist').all()
    serializer_class = InvoiceSettingsSerializer
    permission_classes = [IsAuthenticated, CanManageInvoices]
    http_method_names = ['get', 'post', 'put', 'patch']

    def list(self, request, *args, **kwargs):
        """Get or create invoice settings (singleton)"""
        settings, created = InvoiceSettings.objects.get_or_create()
        serializer = self.get_serializer(settings)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        """Update settings (create not allowed, use update instead)"""
        settings, created = InvoiceSettings.objects.get_or_create()
        serializer = self.get_serializer(settings, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def update(self, request, *args, **kwargs):
        """Update invoice settings"""
        settings, created = InvoiceSettings.objects.get_or_create()
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(settings, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='upload-logo')
    def upload_logo(self, request):
        """Upload company logo"""
        settings, created = InvoiceSettings.objects.get_or_create()

        if 'logo' not in request.FILES:
            return Response(
                {'error': 'No logo file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )

        settings.logo = request.FILES['logo']
        settings.save()

        serializer = self.get_serializer(settings)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='upload-badge')
    def upload_badge(self, request):
        """Upload company badge/seal"""
        settings, created = InvoiceSettings.objects.get_or_create()

        if 'badge' not in request.FILES:
            return Response(
                {'error': 'No badge file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )

        settings.badge = request.FILES['badge']
        settings.save()

        serializer = self.get_serializer(settings)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='upload-signature')
    def upload_signature(self, request):
        """Upload signature image"""
        settings, created = InvoiceSettings.objects.get_or_create()

        if 'signature' not in request.FILES:
            return Response(
                {'error': 'No signature file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )

        settings.signature = request.FILES['signature']
        settings.save()

        serializer = self.get_serializer(settings)
        return Response(serializer.data)

    @action(detail=False, methods=['delete'], url_path='remove-logo')
    def remove_logo(self, request):
        """Remove company logo"""
        settings = InvoiceSettings.objects.first()
        if settings and settings.logo:
            settings.logo.delete()
            settings.save()

        return Response({'message': 'Logo removed successfully'})

    @action(detail=False, methods=['delete'], url_path='remove-badge')
    def remove_badge(self, request):
        """Remove company badge"""
        settings = InvoiceSettings.objects.first()
        if settings and settings.badge:
            settings.badge.delete()
            settings.save()

        return Response({'message': 'Badge removed successfully'})

    @action(detail=False, methods=['delete'], url_path='remove-signature')
    def remove_signature(self, request):
        """Remove signature"""
        settings = InvoiceSettings.objects.first()
        if settings and settings.signature:
            settings.signature.delete()
            settings.save()

        return Response({'message': 'Signature removed successfully'})

    @action(detail=False, methods=['get'], url_path='debug-subscription')
    def debug_subscription(self, request):
        """Debug endpoint to check subscription state"""
        from tenants.permissions import get_tenant_subscription, has_subscription_feature

        subscription = get_tenant_subscription(request)

        if not subscription:
            return Response({
                'error': 'No subscription found',
                'tenant': request.tenant.schema_name if hasattr(request, 'tenant') else 'no-tenant'
            })

        # Get selected features
        selected_features = list(subscription.selected_features.values('id', 'key', 'name'))

        # Test the actual permission function
        has_feature_result = has_subscription_feature(request, 'invoice_management')

        return Response({
            'tenant': request.tenant.schema_name,
            'subscription': {
                'id': subscription.id,
                'is_active': subscription.is_active,
                'subscription_type': subscription.subscription_type,
                'agent_count': subscription.agent_count,
                'monthly_cost': float(subscription.monthly_cost),
            },
            'selected_features': selected_features,
            'has_invoice_management_query': subscription.selected_features.filter(key='invoice_management').exists(),
            'has_invoice_management_function': has_feature_result,
        })

    @action(detail=False, methods=['get'], url_path='available-itemlists')
    def available_itemlists(self, request):
        """Get list of available item lists that can be used for clients"""
        from tickets.models import ItemList
        from tickets.serializers import ItemListMinimalSerializer

        item_lists = ItemList.objects.filter(is_active=True)
        serializer = ItemListMinimalSerializer(item_lists, many=True)
        return Response(serializer.data)


class InvoiceViewSet(viewsets.ModelViewSet):
    """
    API endpoint for invoice management
    """
    permission_classes = [IsAuthenticated, CanManageInvoices]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'client', 'currency']
    search_fields = ['invoice_number', 'client__first_name', 'client__last_name', 'client__email']
    ordering_fields = ['created_at', 'issue_date', 'due_date', 'total']
    ordering = ['-created_at']

    def get_queryset(self):
        return Invoice.objects.select_related(
            'client', 'client_itemlist_item', 'created_by', 'template',
        ).prefetch_related(
            Prefetch(
                'line_items',
                queryset=InvoiceLineItem.objects.select_related('product', 'list_item'),
            ),
            Prefetch(
                'payments',
                queryset=InvoicePayment.objects.select_related('recorded_by'),
            ),
        ).annotate(
            line_items_count=Count('line_items', distinct=True),
        )

    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return InvoiceListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return InvoiceCreateUpdateSerializer
        return InvoiceDetailSerializer

    def list(self, request, *args, **kwargs):
        """List invoices with filters"""
        queryset = self.filter_queryset(self.get_queryset())

        # Additional filters
        status_filter = request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(issue_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(issue_date__lte=date_to)

        # Check for overdue invoices
        overdue = request.query_params.get('overdue')
        if overdue == 'true':
            queryset = queryset.filter(
                due_date__lt=timezone.now().date(),
                status__in=['sent', 'viewed', 'partially_paid']
            )

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        """Get invoice details"""
        return super().retrieve(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        """Create new invoice"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Set created_by
        invoice = serializer.save(created_by=request.user)

        # Return detailed response
        detail_serializer = InvoiceDetailSerializer(invoice, context={'request': request})
        return Response(detail_serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        """Update invoice (only if draft)"""
        instance = self.get_object()

        if instance.status != 'draft':
            return Response(
                {'error': 'Only draft invoices can be edited'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Delete invoice (only if draft)"""
        instance = self.get_object()

        if instance.status != 'draft':
            return Response(
                {'error': 'Only draft invoices can be deleted'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def finalize(self, request, pk=None):
        """Finalize invoice and generate PDF"""
        invoice = self.get_object()

        if invoice.status != 'draft':
            return Response(
                {'error': 'Only draft invoices can be finalized'},
                status=status.HTTP_400_BAD_REQUEST
            )

        invoice.status = 'sent'
        invoice.save()

        # Generate PDF asynchronously via Celery
        from .tasks import generate_invoice_pdf
        schema_name = request.tenant.schema_name
        generate_invoice_pdf.delay(schema_name, invoice.id)

        serializer = self.get_serializer(invoice)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def send_email(self, request, pk=None):
        """Send invoice via email"""
        invoice = self.get_object()

        if invoice.status == 'draft':
            return Response(
                {'error': 'Finalize invoice before sending'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate required fields
        recipient_email = request.data.get('recipient_email')
        if not recipient_email:
            return Response(
                {'error': 'recipient_email is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        cc_emails = request.data.get('cc_emails', [])
        if isinstance(cc_emails, str):
            cc_emails = [e.strip() for e in cc_emails.split(',') if e.strip()]

        subject = request.data.get('subject', '')
        message = request.data.get('message', '')
        attach_pdf = request.data.get('attach_pdf', True)

        # Dispatch Celery task
        from .tasks import send_invoice_email
        schema_name = request.tenant.schema_name
        send_invoice_email.delay(
            schema_name,
            invoice.id,
            recipient_email,
            cc_emails=cc_emails,
            subject=subject,
            message=message,
            attach_pdf=attach_pdf,
        )

        invoice.sent_date = timezone.now()
        invoice.save()

        return Response({'message': 'Invoice email queued for delivery'})

    @action(detail=True, methods=['get'])
    def pdf(self, request, pk=None):
        """Download invoice PDF"""
        invoice = self.get_object()

        # Check if we should regenerate the PDF
        regenerate = request.query_params.get('regenerate', 'false').lower() == 'true'

        # Generate PDF if it doesn't exist or regeneration is requested
        if not invoice.pdf_file or regenerate:
            try:
                invoice.generate_pdf()
            except Exception as e:
                return Response(
                    {'error': f'Failed to generate PDF: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        # Return the PDF file
        if invoice.pdf_file:
            try:
                response = FileResponse(
                    invoice.pdf_file.open('rb'),
                    content_type='application/pdf'
                )
                response['Content-Disposition'] = f'inline; filename="{invoice.invoice_number}.pdf"'
                return response
            except Exception as e:
                return Response(
                    {'error': f'Failed to open PDF file: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        else:
            return Response(
                {'error': 'PDF file not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=True, methods=['get'])
    def excel(self, request, pk=None):
        """Export invoice to Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from io import BytesIO

        invoice = self.get_object()

        try:
            inv_settings = InvoiceSettings.objects.first()
        except InvoiceSettings.DoesNotExist:
            inv_settings = None

        wb = Workbook()
        ws = wb.active
        ws.title = f"Invoice {invoice.invoice_number}"

        # Styles
        bold = Font(bold=True)
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_text = Font(bold=True, color="FFFFFF")

        row = 1

        # Company info
        company_name = inv_settings.company_name if inv_settings else ''
        if company_name:
            ws.cell(row=row, column=1, value=company_name).font = header_font
            row += 1
        if inv_settings and inv_settings.address:
            ws.cell(row=row, column=1, value=inv_settings.address)
            row += 1
        if inv_settings and inv_settings.phone:
            ws.cell(row=row, column=1, value=f"Phone: {inv_settings.phone}")
            row += 1
        if inv_settings and inv_settings.email:
            ws.cell(row=row, column=1, value=f"Email: {inv_settings.email}")
            row += 1
        row += 1

        # Invoice metadata
        ws.cell(row=row, column=1, value="Invoice Number:").font = bold
        ws.cell(row=row, column=2, value=invoice.invoice_number)
        row += 1
        ws.cell(row=row, column=1, value="Status:").font = bold
        ws.cell(row=row, column=2, value=invoice.get_status_display())
        row += 1
        ws.cell(row=row, column=1, value="Issue Date:").font = bold
        ws.cell(row=row, column=2, value=str(invoice.issue_date))
        row += 1
        ws.cell(row=row, column=1, value="Due Date:").font = bold
        ws.cell(row=row, column=2, value=str(invoice.due_date))
        row += 1
        ws.cell(row=row, column=1, value="Currency:").font = bold
        ws.cell(row=row, column=2, value=invoice.currency)
        row += 1

        # Client info
        client_name = invoice.client_name
        if not client_name:
            if invoice.client_itemlist_item:
                client_name = invoice.client_itemlist_item.label
            elif invoice.client:
                client_name = f"{invoice.client.first_name} {invoice.client.last_name}".strip()
        ws.cell(row=row, column=1, value="Client:").font = bold
        ws.cell(row=row, column=2, value=client_name or '')
        row += 2

        # Line items table header
        ws.cell(row=row, column=1, value="Line Items").font = subheader_font
        row += 1

        table_headers = ["#", "Description", "Qty", "Unit", "Unit Price", "Tax %", "Discount %", "Total"]
        for col_idx, header in enumerate(table_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_text
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # Line items
        line_items = invoice.line_items.all().order_by('position')
        for idx, item in enumerate(line_items, 1):
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=item.description).border = thin_border
            ws.cell(row=row, column=3, value=float(item.quantity)).border = thin_border
            ws.cell(row=row, column=4, value=item.unit).border = thin_border
            ws.cell(row=row, column=5, value=float(item.unit_price)).border = thin_border
            ws.cell(row=row, column=6, value=float(item.tax_rate)).border = thin_border
            ws.cell(row=row, column=7, value=float(item.discount_percent)).border = thin_border
            ws.cell(row=row, column=8, value=float(item.line_total)).border = thin_border

            # Right-align number columns
            for col in [3, 5, 6, 7, 8]:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
            row += 1

        row += 1

        # Totals
        ws.cell(row=row, column=7, value="Subtotal:").font = bold
        ws.cell(row=row, column=8, value=float(invoice.subtotal)).font = bold
        row += 1
        ws.cell(row=row, column=7, value="Tax:").font = bold
        ws.cell(row=row, column=8, value=float(invoice.tax_amount)).font = bold
        row += 1
        if invoice.discount_amount:
            ws.cell(row=row, column=7, value="Discount:").font = bold
            ws.cell(row=row, column=8, value=float(-invoice.discount_amount)).font = bold
            row += 1
        ws.cell(row=row, column=7, value="Total:").font = Font(bold=True, size=12)
        ws.cell(row=row, column=8, value=float(invoice.total)).font = Font(bold=True, size=12)
        row += 2

        # Notes
        if invoice.notes:
            ws.cell(row=row, column=1, value="Notes:").font = bold
            row += 1
            ws.cell(row=row, column=1, value=invoice.notes)
            row += 2

        # Terms
        if invoice.terms_and_conditions:
            ws.cell(row=row, column=1, value="Terms & Conditions:").font = bold
            row += 1
            ws.cell(row=row, column=1, value=invoice.terms_and_conditions)

        # Auto-fit column widths
        col_widths = [5, 40, 8, 8, 12, 8, 10, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Write to response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{invoice.invoice_number}.xlsx"'
        return response

    @action(detail=True, methods=['post'])
    def mark_paid(self, request, pk=None):
        """Quick mark invoice as paid"""
        invoice = self.get_object()

        # Create payment record
        InvoicePayment.objects.create(
            invoice=invoice,
            amount=invoice.total - invoice.paid_amount,
            payment_method=request.data.get('payment_method', 'bank_transfer'),
            payment_date=timezone.now(),
            recorded_by=request.user,
            notes=request.data.get('notes', '')
        )

        invoice.mark_as_paid()

        serializer = self.get_serializer(invoice)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Duplicate an existing invoice"""
        original = self.get_object()

        # Create new invoice with copied data
        new_invoice_data = {
            'client': original.client.id,
            'currency': original.currency,
            'discount_amount': original.discount_amount,
            'notes': original.notes,
            'terms_and_conditions': original.terms_and_conditions,
            'template': original.template.id if original.template else None,
            'issue_date': timezone.now().date(),
            'due_date': timezone.now().date() + timedelta(days=30),
        }

        # Copy line items
        line_items = []
        for item in original.line_items.all():
            line_items.append({
                'item_source': item.item_source,
                'product': item.product.id if item.product else None,
                'list_item': item.list_item.id if item.list_item else None,
                'description': item.description,
                'quantity': item.quantity,
                'unit': item.unit,
                'unit_price': item.unit_price,
                'tax_rate': item.tax_rate,
                'discount_percent': item.discount_percent,
                'position': item.position,
            })

        new_invoice_data['line_items'] = line_items

        serializer = InvoiceCreateUpdateSerializer(data=new_invoice_data)
        serializer.is_valid(raise_exception=True)
        new_invoice = serializer.save(created_by=request.user)

        detail_serializer = InvoiceDetailSerializer(new_invoice, context={'request': request})
        return Response(detail_serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get invoice statistics"""
        queryset = self.get_queryset()

        # Current month stats
        now = timezone.now()
        current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        current_month = queryset.filter(created_at__gte=current_month_start)

        total_invoiced = current_month.aggregate(Sum('total'))['total__sum'] or 0
        total_paid = current_month.filter(status='paid').aggregate(Sum('total'))['total__sum'] or 0
        outstanding = queryset.exclude(status__in=['paid', 'cancelled']).aggregate(Sum('total'))['total__sum'] or 0
        overdue_count = queryset.filter(
            due_date__lt=now.date(),
            status__in=['sent', 'viewed', 'partially_paid']
        ).count()

        return Response({
            'current_month': {
                'total_invoiced': total_invoiced,
                'total_paid': total_paid,
                'count': current_month.count()
            },
            'outstanding_amount': outstanding,
            'overdue_count': overdue_count
        })


class InvoiceLineItemViewSet(viewsets.ModelViewSet):
    """
    API endpoint for invoice line items
    """
    queryset = InvoiceLineItem.objects.select_related('product', 'list_item').all()
    serializer_class = InvoiceLineItemSerializer
    permission_classes = [IsAuthenticated, CanManageInvoices]

    def list(self, request, *args, **kwargs):
        """List line items (filtered by invoice if provided)"""
        queryset = self.get_queryset()

        invoice_id = request.query_params.get('invoice_id')
        if invoice_id:
            queryset = queryset.filter(invoice_id=invoice_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='reorder')
    def reorder(self, request):
        """Reorder line items"""
        items = request.data.get('items', [])

        for item_data in items:
            item = InvoiceLineItem.objects.get(id=item_data['id'])
            item.position = item_data['position']
            item.save()

        return Response({'message': 'Line items reordered successfully'})


class InvoicePaymentViewSet(viewsets.ModelViewSet):
    """
    API endpoint for invoice payments
    """
    queryset = InvoicePayment.objects.select_related('recorded_by').all()
    serializer_class = InvoicePaymentSerializer
    permission_classes = [IsAuthenticated, CanManageInvoices]

    def list(self, request, *args, **kwargs):
        """List payments (filtered by invoice if provided)"""
        queryset = self.get_queryset()

        invoice_id = request.query_params.get('invoice_id')
        if invoice_id:
            queryset = queryset.filter(invoice_id=invoice_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        """Record a payment"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(recorded_by=request.user)

        return Response(serializer.data, status=status.HTTP_201_CREATED)


class InvoiceTemplateViewSet(viewsets.ModelViewSet):
    """
    API endpoint for invoice templates
    """
    queryset = InvoiceTemplate.objects.select_related('created_by').filter(is_active=True)
    serializer_class = InvoiceTemplateSerializer
    permission_classes = [IsAuthenticated, CanManageInvoices]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'description']

    def list(self, request, *args, **kwargs):
        """List active templates"""
        return super().list(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        """Create new template"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(created_by=request.user)

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'])
    def preview(self, request, pk=None):
        """Preview template with sample data"""
        template = self.get_object()

        # TODO: Implement template preview with sample data
        return Response(
            {'message': 'Template preview not yet implemented'},
            status=status.HTTP_501_NOT_IMPLEMENTED
        )


class ClientViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for client selection in invoices
    Returns clients from the configured client_itemlist in InvoiceSettings
    """
    serializer_class = ClientSerializer
    permission_classes = [IsAuthenticated, CanManageInvoices]
    filter_backends = [filters.SearchFilter]
    search_fields = ['label']

    def get_queryset(self):
        """
        Get clients from the configured ItemList in InvoiceSettings
        Falls back to EcommerceClient if no ItemList is configured
        """
        import logging
        logger = logging.getLogger(__name__)

        # Get invoice settings to find the client itemlist
        settings, created = InvoiceSettings.objects.get_or_create()

        logger.info(f"[ClientViewSet] Settings: client_itemlist={settings.client_itemlist}")

        if settings.client_itemlist:
            # Return items from the selected ItemList
            queryset = ListItem.objects.filter(
                item_list=settings.client_itemlist,
                is_active=True
            ).select_related('item_list')
            self.search_fields = ['label']
            logger.info(f"[ClientViewSet] Returning ListItem queryset, count={queryset.count()}")
            return queryset
        else:
            # Fallback to EcommerceClient for backward compatibility
            queryset = EcommerceClient.objects.filter(is_active=True)
            self.search_fields = ['first_name', 'last_name', 'email']
            logger.info(f"[ClientViewSet] Returning EcommerceClient queryset, count={queryset.count()}")
            return queryset


class InvoiceMaterialViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for materials (from ItemList) selection
    """
    queryset = ListItem.objects.filter(is_active=True).select_related('item_list')
    serializer_class = ListItemMaterialSerializer
    permission_classes = [IsAuthenticated, CanManageInvoices]
    filter_backends = [filters.SearchFilter]
    search_fields = ['label', 'custom_data']

    def list(self, request, *args, **kwargs):
        """List materials from configured materials ItemList in InvoiceSettings"""
        settings, _ = InvoiceSettings.objects.get_or_create()

        if settings.materials_itemlist:
            queryset = self.get_queryset().filter(item_list=settings.materials_itemlist)
        else:
            queryset = self.get_queryset().none()

        # Apply search
        queryset = self.filter_queryset(queryset)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
